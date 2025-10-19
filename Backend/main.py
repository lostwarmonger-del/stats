#!/usr/bin/env python3
# main.py - Working FastAPI server for Program 404640723

from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import timedelta
from typing import List
import hashlib
from jose import JWTError, jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy import create_engine, Column, Integer, String, Boolean, DateTime, Text, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import io
from fastapi.responses import Response

# Database setup
DATABASE_URL = "sqlite:///./users.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# User model
class User(Base):
    __tablename__ = "users"
    
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    name = Column(String(100), nullable=False)
    hashed_password = Column(String(255), nullable=False)
    role = Column(String(20), default="user", nullable=False)
    is_active = Column(Boolean, default=True)

# Notification model
class Notification(Base):
    __tablename__ = "notifications"
    
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String(200), nullable=False)
    message = Column(String(1000), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    created_by = Column(Integer, nullable=False)  # User ID who created the notification

# Program 404640723 Data Model
class Program404640723Data(Base):
    __tablename__ = "program_404640723_data"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    
    # Form fields - all as strings for Shamsi dates
    row_number = Column(String(10))
    page_channel_address = Column(String(500))
    news_subject = Column(String(200))
    naba_code = Column(String(50))
    naba_code_review = Column(String(100))
    naba_code_registration_date = Column(String(20))
    suspicion_classification = Column(String(100))
    suspicion_registration_date = Column(String(20))
    action_results = Column(Text)
    identified_military_count = Column(String(10))
    identified_military_dependents_count = Column(String(10))
    honorary_police = Column(String(100))
    
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Program 404640725 Data Model
class Program404640725Data(Base):
    __tablename__ = "program_404640725_data"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    
    # Form fields - all as strings for Shamsi dates
    row_number = Column(String(10))
    page_channel_address = Column(String(500))
    news_subject = Column(String(200))
    naba_code = Column(String(50))
    naba_code_review = Column(String(100))
    naba_code_registration_date = Column(String(20))
    suspicion_classification = Column(String(100))
    suspicion_registration_date = Column(String(20))
    action_results = Column(Text)
    identified_military_count = Column(String(10))
    identified_military_dependents_count = Column(String(10))
    honorary_police = Column(String(100))
    
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Program 404640758 Data Model - Personal Information Form
class Program404640758Data(Base):
    __tablename__ = "program_404640758_data"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False, index=True)
    
    # Personal Information
    first_name = Column(String(100))  # نام
    last_name = Column(String(100))  # نام خانوادگی
    father_name = Column(String(100))  # نام پدر
    id_number = Column(String(20))  # شماره شناسنامه
    birth_date = Column(String(20))  # تاریخ تولد (Shamsi)
    national_id = Column(String(20))  # کد ملی
    
    # ID Card Information
    id_issue_date = Column(String(20))  # تاریخ صدور شناسنامه (Shamsi)
    id_issue_place = Column(String(100))  # محل صدور شناسنامه
    religion = Column(String(50))  # دین (مذهب)
    
    # Location Information
    province = Column(String(100))  # استان
    city = Column(String(100))  # شهر
    address = Column(Text)  # آدرس محل سکونت
    phone_number = Column(String(20))  # شماره تماس
    
    # Personal Status
    marital_status = Column(String(50))  # وضعیت تأهل
    military_service_status = Column(String(50))  # وضعیت خدمت وظیفه
    physical_status = Column(String(50))  # وضعیت جسمانی
    health_description = Column(Text)  # توضیح سطح سلامت
    nationality = Column(String(50))  # ملیت
    
    # Professional Information
    activity_field = Column(String(100))  # حوزه فعالیت
    educational_background = Column(Text)  # سوابق تحصیلی
    work_experience = Column(Text)  # تجربیات شغلی
    foreign_languages = Column(Text)  # زبان های خارجی
    certificates = Column(Text)  # گواهینامه دوره های گذرانده شده
    requested_cooperation = Column(Text)  # همکاری پیشنهادی مورد درخواست
    proposed_salary = Column(String(20))  # حقوق پیشنهادی
    scientific_activities = Column(Text)  # فعالیت های علمی
    
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Create tables
def create_tables():
    Base.metadata.create_all(bind=engine)

# Database dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Create default admin user
def create_default_admin():
    db = SessionLocal()
    try:
        admin_user = db.query(User).filter(User.username == "admin").first()
        if not admin_user:
            hashed_password = get_password_hash("admin123")
            admin_user = User(
                username="admin",
                name="مدیر سیستم",
                hashed_password=hashed_password,
                role="admin",
                is_active=True
            )
            db.add(admin_user)
            db.commit()
            print("Default admin user created: admin/admin123")
    finally:
        db.close()

# Pydantic models
class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    user: dict

class NotificationCreateRequest(BaseModel):
    title: str
    message: str

class NotificationResponse(BaseModel):
    id: int
    title: str
    message: str
    created_at: datetime
    created_by: int
    created_by_name: str

class Program404640723Request(BaseModel):
    row_number: str
    page_channel_address: str
    news_subject: str
    naba_code: int  # Changed to int
    naba_code_review: str
    naba_code_registration_date: str
    suspicion_classification: int  # Changed to int
    suspicion_registration_date: str
    action_results: str
    identified_military_count: int  # Changed to int
    identified_military_dependents_count: int  # Changed to int
    honorary_police: int  # Changed to int

class Program404640725Request(BaseModel):
    row_number: str
    page_channel_address: str
    news_subject: str
    naba_code: int  # Changed to int
    naba_code_review: str
    naba_code_registration_date: str
    suspicion_classification: int  # Changed to int
    suspicion_registration_date: str
    action_results: str
    identified_military_count: int  # Changed to int
    identified_military_dependents_count: int  # Changed to int
    honorary_police: int  # Changed to int

class Program404640758Request(BaseModel):
    first_name: str = ""
    last_name: str = ""
    father_name: str = ""
    id_number: str = ""
    birth_date: str = ""  # Shamsi date as string
    national_id: str = ""
    id_issue_date: str = ""  # Shamsi date as string
    id_issue_place: str = ""
    religion: str = ""
    province: str = ""
    city: str = ""
    marital_status: str = ""
    military_service_status: str = ""
    physical_status: str = ""
    health_description: str = ""
    nationality: str = ""
    activity_field: str = ""
    educational_background: str = ""
    work_experience: str = ""
    foreign_languages: str = ""
    certificates: str = ""
    requested_cooperation: str = ""
    proposed_salary: str = ""
    scientific_activities: str = ""
    address: str = ""
    phone_number: str = ""
    
    class Config:
        # Allow extra fields and convert None to empty string
        extra = "ignore"
        validate_assignment = True

# Auth configuration
SECRET_KEY = "your-secret-key-change-this-in-production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Password hashing
def verify_password(plain_password, hashed_password):
    return hashlib.sha256(plain_password.encode()).hexdigest() == hashed_password

def get_password_hash(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Security
security = HTTPBearer()

def authenticate_user(db: Session, username: str, password: str):
    user = db.query(User).filter(User.username == username).first()
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    return user

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise credentials_exception
    return user

def get_current_active_user(current_user: User = Depends(get_current_user)):
    if not current_user.is_active:
        raise HTTPException(status_code=400, detail="Inactive user")
    return current_user

# Create FastAPI app
app = FastAPI(
    title="Persian Management System",
    description="A luxury Persian management system",
    version="1.0.0"
)

# CORS middleware - Fixed for your frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for now
    allow_credentials=False,  # Set to False to avoid CORS issues
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)

# Create tables and default admin on startup
@app.on_event("startup")
async def startup_event():
    print("Starting Persian Management System...")
    create_tables()
    create_default_admin()
    print("Server startup complete!")

# Root endpoint
@app.get("/")
async def root():
    return {"message": "Persian Management System API", "version": "1.0.0"}

# Health check endpoint
@app.get("/health")
async def health_check():
    return {"status": "healthy", "message": "API is running"}

# Authentication endpoints
@app.post("/auth/login", response_model=LoginResponse)
async def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    user = authenticate_user(db, login_data.username, login_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="نام کاربری یا رمز عبور اشتباه است",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=30)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    
    user_data = {
        "id": user.id,
        "username": user.username,
        "name": user.name,
        "role": user.role,
        "is_active": user.is_active
    }
    
    return LoginResponse(
        access_token=access_token,
        token_type="bearer",
        user=user_data
    )

@app.get("/auth/me")
async def read_users_me(current_user: User = Depends(get_current_active_user)):
    user_data = {
        "id": current_user.id,
        "username": current_user.username,
        "name": current_user.name,
        "role": current_user.role,
        "is_active": current_user.is_active
    }
    return user_data

# Notification endpoints
@app.get("/notifications")
async def get_notifications(current_user: User = Depends(get_current_active_user)):
    """Get the last 5 notifications with recent notification indicator"""
    db = SessionLocal()
    try:
        notifications = db.query(Notification).order_by(Notification.created_at.desc()).limit(5).all()
        
        result = []
        for notification in notifications:
            # Get the creator's name
            creator = db.query(User).filter(User.id == notification.created_by).first()
            creator_name = creator.name if creator else "کاربر ناشناس"
            
            result.append(NotificationResponse(
                id=notification.id,
                title=notification.title,
                message=notification.message,
                created_at=notification.created_at,
                created_by=notification.created_by,
                created_by_name=creator_name
            ))
        
        # Check if there are any notifications from the last 10 days
        from datetime import datetime, timedelta
        ten_days_ago = datetime.utcnow() - timedelta(days=10)
        recent_notifications = db.query(Notification).filter(Notification.created_at >= ten_days_ago).count()
        has_recent_notifications = recent_notifications > 0
        
        return {
            "notifications": result,
            "has_recent_notifications": has_recent_notifications,
            "recent_count": recent_notifications
        }
    finally:
        db.close()

@app.post("/notifications", response_model=NotificationResponse)
async def create_notification(notification_data: NotificationCreateRequest, current_user: User = Depends(get_current_active_user)):
    """Create a new notification - only admin (user ID 1) can create notifications"""
    if current_user.id != 1:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="فقط مدیر اصلی سیستم می‌تواند اعلان ارسال کند"
        )
    
    db = SessionLocal()
    try:
        new_notification = Notification(
            title=notification_data.title,
            message=notification_data.message,
            created_by=current_user.id
        )
        
        db.add(new_notification)
        db.commit()
        db.refresh(new_notification)
        
        return NotificationResponse(
            id=new_notification.id,
            title=new_notification.title,
            message=new_notification.message,
            created_at=new_notification.created_at,
            created_by=new_notification.created_by,
            created_by_name=current_user.name
        )
    finally:
        db.close()

@app.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: int, current_user: User = Depends(get_current_active_user)):
    """Delete a notification - only admin (user ID 1) can delete notifications"""
    if current_user.id != 1:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="فقط مدیر اصلی سیستم می‌تواند اعلان‌ها را حذف کند"
        )
    
    db = SessionLocal()
    try:
        notification = db.query(Notification).filter(Notification.id == notification_id).first()
        if not notification:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="اعلان یافت نشد"
            )
        
        db.delete(notification)
        db.commit()
        
        return {"message": "اعلان با موفقیت حذف شد"}
    finally:
        db.close()


# Users management endpoints
@app.get("/users")
async def get_users(current_user: User = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        # Include all users including user ID 1 (main admin)
        users = db.query(User).all()
        return users
    finally:
        db.close()

@app.post("/users")
async def create_user(user_data: dict, current_user: User = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        # Check if username already exists
        existing_user = db.query(User).filter(User.username == user_data["username"]).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Username already exists")
        
        new_user = User(
            username=user_data["username"],
            name=user_data["name"],
            hashed_password=get_password_hash(user_data["password"]),
            role=user_data.get("role", "user"),
            is_active=True
        )
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        return {"message": "User created successfully", "user_id": new_user.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating user: {str(e)}")
    finally:
        db.close()

@app.put("/users/{user_id}")
async def update_user(user_id: int, user_data: dict, current_user: User = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent modification of user ID 1 (main admin)
    if user_id == 1:
        raise HTTPException(status_code=403, detail="Cannot modify main admin account")
    
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        if "name" in user_data:
            user.name = user_data["name"]
        if "role" in user_data:
            user.role = user_data["role"]
        if "is_active" in user_data:
            user.is_active = user_data["is_active"]
        if "password" in user_data and user_data["password"]:
            user.hashed_password = get_password_hash(user_data["password"])
        
        db.commit()
        return {"message": "User updated successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating user: {str(e)}")
    finally:
        db.close()

@app.delete("/users/{user_id}")
async def delete_user(user_id: int, current_user: User = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent deletion of user ID 1 (main admin)
    if user_id == 1:
        raise HTTPException(status_code=403, detail="Cannot delete main admin account")
    
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        db.delete(user)
        db.commit()
        return {"message": "User deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting user: {str(e)}")
    finally:
        db.close()

@app.patch("/users/{user_id}/edit")
async def edit_user(user_id: int, user_data: dict, current_user: User = Depends(get_current_active_user)):
    """Edit user name and password - only admins and user ID 1 can edit"""
    # Allow user ID 1 to edit themselves, or admins to edit others
    if current_user.id != 1 and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Prevent user ID 1 from editing other users (only admins can edit others)
    if current_user.id == 1 and user_id != 1 and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Cannot edit other users")
    
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Update name if provided
        if "name" in user_data and user_data["name"]:
            user.name = user_data["name"]
        
        # Update password if provided
        if "password" in user_data and user_data["password"]:
            user.hashed_password = get_password_hash(user_data["password"])
        
        # Update role if provided (only admins can change roles)
        if "role" in user_data and user_data["role"] and current_user.role == "admin":
            user.role = user_data["role"]
        
        db.commit()
        return {"message": "User updated successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating user: {str(e)}")
    finally:
        db.close()

@app.patch("/users/{user_id}/toggle-status")
async def toggle_user_status(user_id: int, current_user: User = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent modification of user ID 1 (main admin)
    if user_id == 1:
        raise HTTPException(status_code=403, detail="Cannot modify main admin account")
    
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        user.is_active = not user.is_active
        db.commit()
        return {"message": f"User {'activated' if user.is_active else 'deactivated'} successfully", "is_active": user.is_active}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error toggling user status: {str(e)}")
    finally:
        db.close()

# Program 404640723 endpoints
@app.get("/current-date")
async def get_current_date():
    """Get current Persian date"""
    from datetime import datetime
    import jdatetime
    
    # Get current Persian date
    now = jdatetime.datetime.now()
    persian_date = f"{now.year}/{now.month:02d}/{now.day:02d}"
    
    return {
        "persian_date": persian_date,
        "gregorian_date": datetime.now().isoformat(),
        "year": now.year,
        "month": now.month,
        "day": now.day
    }

@app.get("/programs/404640723/all")
async def get_all_program_404640723_data(
    user_id: int = None, 
    current_user: User = Depends(get_current_active_user)
):
    print(f"User {current_user.username} requesting program data")
    
    db = SessionLocal()
    try:
        if current_user.role == "admin":
            print("Admin requesting all program data...")
            query = db.query(Program404640723Data)
            
            # If user_id filter is provided, filter by that user
            if user_id is not None:
                query = query.filter(Program404640723Data.user_id == user_id)
                print(f"Filtering by user_id: {user_id}")
            
            data = query.order_by(Program404640723Data.created_at.desc()).all()
        else:
            print(f"Regular user requesting their own data...")
            data = db.query(Program404640723Data).filter(Program404640723Data.user_id == current_user.id).order_by(Program404640723Data.created_at.desc()).all()
        
        print(f"Found {len(data)} records")
        
        # Convert SQLAlchemy objects to dictionaries for JSON serialization
        result = []
        for item in data:
            # Get user name for this record
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            result.append({
                "id": item.id,
                "user_id": item.user_id,
                "user_name": user_name,
                "row_number": item.row_number,
                "page_channel_address": item.page_channel_address,
                "news_subject": item.news_subject,
                "naba_code": item.naba_code,
                "naba_code_review": item.naba_code_review,
                "naba_code_registration_date": item.naba_code_registration_date,
                "suspicion_classification": item.suspicion_classification,
                "suspicion_registration_date": item.suspicion_registration_date,
                "action_results": item.action_results,
                "identified_military_count": item.identified_military_count,
                "identified_military_dependents_count": item.identified_military_dependents_count,
                "honorary_police": item.honorary_police,
                "created_at": item.created_at,
                "updated_at": item.updated_at
            })
        print(f"Returning {len(result)} records")
        return result
    except Exception as e:
        print(f"Error in get_all_program_404640723_data: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640723/export")
async def export_program_data(current_user: User = Depends(get_current_active_user)):
    """Export program data to CSV/Excel"""
    db = SessionLocal()
    try:
        # Admins can export all data, regular users can only export their own data
        if current_user.role == "admin":
            data = db.query(Program404640723Data).order_by(Program404640723Data.created_at.desc()).all()
        else:
            data = db.query(Program404640723Data).filter(Program404640723Data.user_id == current_user.id).order_by(Program404640723Data.created_at.desc()).all()
        
        # Prepare data for export
        export_data = []
        for item in data:
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            export_data.append({
                "شماره ردیف": item.row_number or "",
                "نام کاربر": user_name,
                "آدرس صفحه/کانال": item.page_channel_address or "",
                "موضوع خبر": item.news_subject or "",
                "کد نبا": item.naba_code or "",
                "بررسی کد نبا": item.naba_code_review or "",
                "تاریخ ثبت کد نبا": item.naba_code_registration_date or "",
                "کلاسه مظنونیت": item.suspicion_classification or "",
                "تاریخ ثبت کلاسه مظنونیت": item.suspicion_registration_date or "",
                "نتایج اقدامات": item.action_results or "",
                "تعداد نظامی شناسایی شده": item.identified_military_count or "",
                "تعداد وابستگان نظامی": item.identified_military_dependents_count or "",
                "پلیس افتخاری": item.honorary_police or "",
                "تاریخ ایجاد": item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else "",
                "آخرین بروزرسانی": item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else ""
            })
        
        # Convert to Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "داده‌های برنامه 404640723"
        
        # Set RTL direction for the worksheet
        ws.sheet_view.rightToLeft = True
        
        if export_data:
            # Headers
            headers = list(export_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row, data_row in enumerate(export_data, 2):
                for col, value in enumerate(data_row.values(), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = str(ws[f"{column_letter}{row}"].value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return Excel file
        from fastapi.responses import Response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=program_404640723_data.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")
    finally:
        db.close()

@app.delete("/programs/404640723/{record_id}")
async def delete_program_404640723_data(record_id: int, current_user: User = Depends(get_current_active_user)):
    """Delete a program data record (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        # Find the record
        record = db.query(Program404640723Data).filter(Program404640723Data.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        
        # Delete the record
        db.delete(record)
        db.commit()
        
        return {"message": "Record deleted successfully", "success": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting record: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640723/{user_id}")
async def get_program_404640723_data(user_id: int, current_user: User = Depends(get_current_active_user)):
    if current_user.id != user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Access denied")
    
    db = SessionLocal()
    try:
        data = db.query(Program404640723Data).filter(Program404640723Data.user_id == user_id).first()
        if not data:
            raise HTTPException(status_code=404, detail="No data found for this user")
        
        # Convert SQLAlchemy object to dictionary for JSON serialization
        return {
            "id": data.id,
            "user_id": data.user_id,
            "row_number": data.row_number,
            "page_channel_address": data.page_channel_address,
            "news_subject": data.news_subject,
            "naba_code": data.naba_code,
            "naba_code_review": data.naba_code_review,
            "naba_code_registration_date": data.naba_code_registration_date,
            "suspicion_classification": data.suspicion_classification,
            "suspicion_registration_date": data.suspicion_registration_date,
            "action_results": data.action_results,
            "identified_military_count": data.identified_military_count,
            "identified_military_dependents_count": data.identified_military_dependents_count,
            "honorary_police": data.honorary_police,
            "created_at": data.created_at,
            "updated_at": data.updated_at
        }
    finally:
        db.close()

@app.post("/programs/404640723")
async def save_program_404640723_data(
    form_data: Program404640723Request, 
    current_user: User = Depends(get_current_active_user)
):
    db = SessionLocal()
    try:
        # Convert form data to dict and convert number fields to strings
        data_dict = form_data.model_dump()
        
        # Convert number fields to strings for database storage
        number_fields = ['naba_code', 'suspicion_classification', 'identified_military_count', 'identified_military_dependents_count', 'honorary_police']
        for field in number_fields:
            if field in data_dict and data_dict[field] is not None:
                data_dict[field] = str(data_dict[field])
        
        # Always create a new record - don't update existing ones
        new_data = Program404640723Data(
            user_id=current_user.id,
            **data_dict
        )
        db.add(new_data)
        db.commit()
        return {"message": "Program 404640723 data saved successfully", "success": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error saving data: {str(e)}")
    finally:
        db.close()

# Program 404640725 endpoints
@app.get("/programs/404640725/all")
async def get_all_program_404640725_data(
    user_id: int = None, 
    current_user: User = Depends(get_current_active_user)
):
    print(f"User {current_user.username} requesting program 404640725 data")
    
    db = SessionLocal()
    try:
        if current_user.role == "admin":
            print("Admin requesting all program 404640725 data...")
            query = db.query(Program404640725Data)
            
            # If user_id filter is provided, filter by that user
            if user_id is not None:
                query = query.filter(Program404640725Data.user_id == user_id)
                print(f"Filtering by user_id: {user_id}")
            
            data = query.order_by(Program404640725Data.created_at.desc()).all()
        else:
            print(f"Regular user requesting their own program 404640725 data...")
            data = db.query(Program404640725Data).filter(Program404640725Data.user_id == current_user.id).order_by(Program404640725Data.created_at.desc()).all()
        
        print(f"Found {len(data)} records")
        
        # Convert SQLAlchemy objects to dictionaries for JSON serialization
        result = []
        for item in data:
            # Get user name for this record
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            result.append({
                "id": item.id,
                "user_id": item.user_id,
                "user_name": user_name,
                "row_number": item.row_number,
                "page_channel_address": item.page_channel_address,
                "news_subject": item.news_subject,
                "naba_code": item.naba_code,
                "naba_code_review": item.naba_code_review,
                "naba_code_registration_date": item.naba_code_registration_date,
                "suspicion_classification": item.suspicion_classification,
                "suspicion_registration_date": item.suspicion_registration_date,
                "action_results": item.action_results,
                "identified_military_count": item.identified_military_count,
                "identified_military_dependents_count": item.identified_military_dependents_count,
                "honorary_police": item.honorary_police,
                "created_at": item.created_at.isoformat(),
                "updated_at": item.updated_at.isoformat()
            })
        
        print(f"Returning {len(result)} records")
        return result
    except Exception as e:
        print(f"Error in get_all_program_404640725_data: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640725/export")
async def export_program_404640725_data(current_user: User = Depends(get_current_active_user)):
    """Export program 404640725 data to CSV/Excel"""
    db = SessionLocal()
    try:
        # Admins can export all data, regular users can only export their own data
        if current_user.role == "admin":
            data = db.query(Program404640725Data).order_by(Program404640725Data.created_at.desc()).all()
        else:
            data = db.query(Program404640725Data).filter(Program404640725Data.user_id == current_user.id).order_by(Program404640725Data.created_at.desc()).all()
        
        # Prepare data for export
        export_data = []
        for item in data:
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            export_data.append({
                "شماره ردیف": item.row_number or "",
                "نام کاربر": user_name,
                "آدرس صفحه/کانال": item.page_channel_address or "",
                "موضوع خبر": item.news_subject or "",
                "کد نبا": item.naba_code or "",
                "بررسی کد نبا": item.naba_code_review or "",
                "تاریخ ثبت کد نبا": item.naba_code_registration_date or "",
                "کلاسه مظنونیت": item.suspicion_classification or "",
                "تاریخ ثبت کلاسه مظنونیت": item.suspicion_registration_date or "",
                "نتایج اقدامات": item.action_results or "",
                "تعداد نظامی شناسایی شده": item.identified_military_count or "",
                "تعداد وابستگان نظامی": item.identified_military_dependents_count or "",
                "پلیس افتخاری": item.honorary_police or "",
                "تاریخ ایجاد": item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else "",
                "آخرین بروزرسانی": item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else ""
            })
        
        # Convert to Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "داده‌های برنامه 404640725"
        
        # Set RTL direction for the worksheet
        ws.sheet_view.rightToLeft = True
        
        if export_data:
            # Headers
            headers = list(export_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row, data_row in enumerate(export_data, 2):
                for col, value in enumerate(data_row.values(), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = str(ws[f"{column_letter}{row}"].value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return Excel file
        from fastapi.responses import Response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=program_404640725_data.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")
    finally:
        db.close()

@app.delete("/programs/404640725/{record_id}")
async def delete_program_404640725_data(record_id: int, current_user: User = Depends(get_current_active_user)):
    """Delete program 404640725 data record (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        record = db.query(Program404640725Data).filter(Program404640725Data.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        
        db.delete(record)
        db.commit()
        return {"message": "Record deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting record: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640725/{user_id}")
async def get_program_404640725_data_by_user(user_id: int, current_user: User = Depends(get_current_active_user)):
    """Get program 404640725 data for a specific user"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        data = db.query(Program404640725Data).filter(Program404640725Data.user_id == user_id).order_by(Program404640725Data.created_at.desc()).all()
        
        result = []
        for item in data:
            result.append({
                "id": item.id,
                "user_id": item.user_id,
                "row_number": item.row_number,
                "page_channel_address": item.page_channel_address,
                "news_subject": item.news_subject,
                "naba_code": item.naba_code,
                "naba_code_review": item.naba_code_review,
                "naba_code_registration_date": item.naba_code_registration_date,
                "suspicion_classification": item.suspicion_classification,
                "suspicion_registration_date": item.suspicion_registration_date,
                "action_results": item.action_results,
                "identified_military_count": item.identified_military_count,
                "identified_military_dependents_count": item.identified_military_dependents_count,
                "honorary_police": item.honorary_police,
                "created_at": item.created_at.isoformat(),
                "updated_at": item.updated_at.isoformat()
            })
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        db.close()

@app.post("/programs/404640725")
async def save_program_404640725_data(
    data: Program404640725Request,
    current_user: User = Depends(get_current_active_user)
):
    """Save program 404640725 data"""
    db = SessionLocal()
    try:
        # Convert int fields to strings for database storage
        data_dict = data.dict()
        data_dict['naba_code'] = str(data.naba_code)
        data_dict['suspicion_classification'] = str(data.suspicion_classification)
        data_dict['identified_military_count'] = str(data.identified_military_count)
        data_dict['identified_military_dependents_count'] = str(data.identified_military_dependents_count)
        data_dict['honorary_police'] = str(data.honorary_police)
        
        new_data = Program404640725Data(
            user_id=current_user.id,
            **data_dict
        )
        db.add(new_data)
        db.commit()
        return {"message": "Program 404640725 data saved successfully", "success": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error saving data: {str(e)}")
    finally:
        db.close()

# Program 404640758 endpoints
@app.get("/programs/404640758/all")
async def get_all_program_404640758_data(
    user_id: int = None, 
    current_user: User = Depends(get_current_active_user)
):
    print(f"User {current_user.username} requesting program 404640758 data")
    
    db = SessionLocal()
    try:
        if current_user.role == "admin":
            print("Admin requesting all program 404640758 data...")
            query = db.query(Program404640758Data)
            
            # If user_id filter is provided, filter by that user
            if user_id is not None:
                query = query.filter(Program404640758Data.user_id == user_id)
                print(f"Filtering by user_id: {user_id}")
            
            data = query.order_by(Program404640758Data.created_at.desc()).all()
        else:
            print(f"Regular user requesting their own program 404640758 data...")
            data = db.query(Program404640758Data).filter(Program404640758Data.user_id == current_user.id).order_by(Program404640758Data.created_at.desc()).all()
        
        print(f"Found {len(data)} records")
        
        # Convert SQLAlchemy objects to dictionaries for JSON serialization
        result = []
        for item in data:
            # Get user name for this record
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            result.append({
                "id": item.id,
                "user_id": item.user_id,
                "user_name": user_name,
                "first_name": item.first_name,
                "last_name": item.last_name,
                "father_name": item.father_name,
                "id_number": item.id_number,
                "birth_date": item.birth_date,
                "national_id": item.national_id,
                "id_issue_date": item.id_issue_date,
                "id_issue_place": item.id_issue_place,
                "religion": item.religion,
                "province": item.province,
                "city": item.city,
                "marital_status": item.marital_status,
                "military_service_status": item.military_service_status,
                "physical_status": item.physical_status,
                "health_description": item.health_description,
                "nationality": item.nationality,
                "activity_field": item.activity_field,
                "educational_background": item.educational_background,
                "work_experience": item.work_experience,
                "foreign_languages": item.foreign_languages,
                "certificates": item.certificates,
                "requested_cooperation": item.requested_cooperation,
                "proposed_salary": item.proposed_salary,
                "scientific_activities": item.scientific_activities,
                "address": item.address,
                "phone_number": item.phone_number,
                "created_at": item.created_at.isoformat(),
                "updated_at": item.updated_at.isoformat()
            })
        
        print(f"Returning {len(result)} records")
        return result
    except Exception as e:
        print(f"Error in get_all_program_404640758_data: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640758/export")
async def export_program_404640758_data(current_user: User = Depends(get_current_active_user)):
    """Export program 404640758 data to Excel"""
    db = SessionLocal()
    try:
        # Admins can export all data, regular users can only export their own data
        if current_user.role == "admin":
            data = db.query(Program404640758Data).order_by(Program404640758Data.created_at.desc()).all()
        else:
            data = db.query(Program404640758Data).filter(Program404640758Data.user_id == current_user.id).order_by(Program404640758Data.created_at.desc()).all()
        
        # Prepare data for export
        export_data = []
        for item in data:
            user = db.query(User).filter(User.id == item.user_id).first()
            user_name = user.name if user else "کاربر ناشناس"
            
            export_data.append({
                "نام کاربر": user_name,
                "نام": item.first_name or "",
                "نام خانوادگی": item.last_name or "",
                "نام پدر": item.father_name or "",
                "شماره شناسنامه": item.id_number or "",
                "تاریخ تولد": item.birth_date or "",
                "کد ملی": item.national_id or "",
                "تاریخ صدور شناسنامه": item.id_issue_date or "",
                "محل صدور شناسنامه": item.id_issue_place or "",
                "دین (مذهب)": item.religion or "",
                "استان": item.province or "",
                "شهر": item.city or "",
                "آدرس محل سکونت": item.address or "",
                "شماره تماس": item.phone_number or "",
                "وضعیت تأهل": item.marital_status or "",
                "وضعیت خدمت وظیفه": item.military_service_status or "",
                "وضعیت جسمانی": item.physical_status or "",
                "توضیح سطح سلامت": item.health_description or "",
                "ملیت": item.nationality or "",
                "حوزه فعالیت": item.activity_field or "",
                "سوابق تحصیلی": item.educational_background or "",
                "تجربیات شغلی": item.work_experience or "",
                "زبان های خارجی": item.foreign_languages or "",
                "گواهینامه دوره های گذرانده شده": item.certificates or "",
                "همکاری پیشنهادی مورد درخواست": item.requested_cooperation or "",
                "حقوق پیشنهادی": item.proposed_salary or "",
                "فعالیت های علمی": item.scientific_activities or "",
                "تاریخ ایجاد": item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else "",
                "آخرین بروزرسانی": item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else ""
            })
        
        # Convert to Excel
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "داده‌های برنامه 404640758"
        
        # Set RTL direction for the worksheet
        ws.sheet_view.rightToLeft = True
        
        if export_data:
            # Headers
            headers = list(export_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row, data_row in enumerate(export_data, 2):
                for col, value in enumerate(data_row.values(), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = str(ws[f"{column_letter}{row}"].value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return Excel file
        from fastapi.responses import Response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=program_404640758_data.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")
    finally:
        db.close()

@app.get("/programs/404640758/{user_id}")
async def get_program_404640758_data(user_id: int, current_user: User = Depends(get_current_active_user)):
    """Get program 404640758 data for a specific user"""
    if current_user.id != user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Access denied")
    
    db = SessionLocal()
    try:
        data = db.query(Program404640758Data).filter(Program404640758Data.user_id == user_id).first()
        if not data:
            raise HTTPException(status_code=404, detail="No data found for this user")
        
        return {
            "id": data.id,
            "user_id": data.user_id,
            "first_name": data.first_name,
            "last_name": data.last_name,
            "father_name": data.father_name,
            "id_number": data.id_number,
            "birth_date": data.birth_date,
            "national_id": data.national_id,
            "id_issue_date": data.id_issue_date,
            "id_issue_place": data.id_issue_place,
            "religion": data.religion,
            "province": data.province,
            "city": data.city,
            "marital_status": data.marital_status,
            "military_service_status": data.military_service_status,
            "physical_status": data.physical_status,
            "health_description": data.health_description,
            "nationality": data.nationality,
            "activity_field": data.activity_field,
            "educational_background": data.educational_background,
            "work_experience": data.work_experience,
            "foreign_languages": data.foreign_languages,
            "certificates": data.certificates,
            "requested_cooperation": data.requested_cooperation,
            "proposed_salary": data.proposed_salary,
            "scientific_activities": data.scientific_activities,
            "address": data.address,
            "phone_number": data.phone_number,
            "created_at": data.created_at,
            "updated_at": data.updated_at
        }
    finally:
        db.close()

@app.post("/programs/404640758")
async def save_program_404640758_data(
    form_data: Program404640758Request, 
    current_user: User = Depends(get_current_active_user)
):
    """Save or update program 404640758 data"""
    db = SessionLocal()
    try:
        # Convert form data to dict and handle None values
        data_dict = form_data.dict()
        
        # Convert None values to empty strings and ensure all values are strings
        for key, value in data_dict.items():
            if value is None:
                data_dict[key] = ""
            else:
                data_dict[key] = str(value)
        
        # Check if user already has data
        existing = db.query(Program404640758Data).filter(
            Program404640758Data.user_id == current_user.id
        ).first()
        
        if existing:
            # Update existing data
            for field, value in data_dict.items():
                if hasattr(existing, field):
                    setattr(existing, field, value)
            existing.updated_at = datetime.utcnow()
        else:
            # Create new data
            new_data = Program404640758Data(
                user_id=current_user.id,
                **data_dict
            )
            db.add(new_data)
        
        db.commit()
        return {"message": "Program 404640758 data saved successfully", "success": True}
    except Exception as e:
        db.rollback()
        print(f"Error saving program 404640758 data: {e}")
        print(f"Form data received: {form_data.dict()}")
        raise HTTPException(status_code=500, detail=f"Error saving data: {str(e)}")
    finally:
        db.close()

@app.delete("/programs/404640758/{record_id}")
async def delete_program_404640758_data(record_id: int, current_user: User = Depends(get_current_active_user)):
    """Delete program 404640758 data record (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = SessionLocal()
    try:
        record = db.query(Program404640758Data).filter(Program404640758Data.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        
        db.delete(record)
        db.commit()
        return {"message": "Record deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting record: {str(e)}")
    finally:
        db.close()

if __name__ == "__main__":
    import uvicorn
    print("Starting server on http://localhost:8001")
    uvicorn.run(app, host="0.0.0.0", port=8001, reload=False)